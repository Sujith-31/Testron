# #Groq API Script V1
# from flask import Flask, render_template, request, Response, stream_with_context, jsonify
# import json
# import io
# import os
# from groq import Groq

# # ── Load .env file ─────────────────────────────────────────────
# try:
#     from dotenv import load_dotenv
#     load_dotenv()
# except ImportError:
#     pass

# try:
#     import openpyxl
#     EXCEL_SUPPORT = True
# except ImportError:
#     EXCEL_SUPPORT = False

# app = Flask(__name__)

# # ── CONFIG ─────────────────────────────────────────────────────
# GROQ_API_KEY = os.environ.get("GROQ_API_KEY")
# MODEL        = os.environ.get("MODEL", "llama-3.3-70b-versatile")
# client       = Groq(api_key=GROQ_API_KEY)

# # ── System Prompt ──────────────────────────────────────────────
# TESTRON_PROMPT = """
# You are Testron — an AI QA Test Case Quality Scoring Engine.

# Your job is to evaluate SOFTWARE TEST CASE QUALITY.

# INPUT:
# You will receive a batch of test cases summarized in compact form.

# CRITICAL RULES:
# - NEVER repeat or quote test cases
# - NEVER summarize the batch
# - NEVER explain the test cases
# - DO NOT write paragraphs
# - ONLY produce the structured QA evaluation

# Treat test case input as CONFIDENTIAL.

# ------------------------------------------------

# Evaluate using professional QA metrics:

# QUALITY CHECK
# + Title clarity
# + Preconditions present
# + Steps completeness
# + Expected result clarity
# + Test data presence

# TEST DESIGN COVERAGE
# + Positive scenarios
# + Negative scenarios
# + Boundary Value Analysis (BVA)
# + Equivalence Partitioning (EP)
# + Edge cases

# ------------------------------------------------

# You must also perform:

# 1. Detect duplicate test cases
# 2. Detect missing scenarios
# 3. Suggest Boundary Value test cases
# 4. Suggest Negative test cases
# 5. Score overall coverage

# ------------------------------------------------

# SCORING MODEL (MAX 10)

# +1 Title clarity
# +1 Preconditions
# +1 Steps completeness
# +1 Expected result
# +1 Test data
# +1 Positive scenarios
# +1 Negative scenarios
# +1 Boundary value tests
# +1 Equivalence partition tests
# +1 Edge cases

# ------------------------------------------------

# Respond ONLY in this format:

# BATCH VERDICT: [X PASS | Y NEEDS FIX | Z CRITICAL] out of [total]

# COMMON ISSUES:
# - issue
# - issue

# DUPLICATE TEST CASES:
# - TC-ID reason
# - TC-ID reason

# COVERAGE GAPS:
# - missing technique
# - missing scenario type

# MISSING TEST CASES:
# - missing scenario
# - missing scenario

# SUGGESTED BVA TESTS:
# 1. boundary case
# 2. boundary case

# SUGGESTED NEGATIVE TESTS:
# 1. negative scenario
# 2. negative scenario

# OVERALL SCORE: X/10

# TOP 3 FIXES:
# 1. improvement
# 2. improvement
# 3. improvement

# IMPORTANT:
# If information is missing, infer what is likely missing based on QA best practices.
# DO NOT output explanations.
# """


# # ── Home Page ──────────────────────────────────────────────────
# @app.route("/")
# def index():
#     return render_template("index.html")


# # ── Chat Endpoint ──────────────────────────────────────────────
# @app.route("/chat", methods=["POST"])
# def chat():
#     data     = request.get_json()
#     messages = data.get("messages", [])

#     # ── Smart System Prompt based on message content ──
#     last_message = messages[-1]["content"].lower() if messages else ""

#     # Detect if it's a casual/general question
#     casual_keywords = [
#         "what is", "what's", "explain", "define", "tell me",
#         "who are you", "your name", "whats your", "hello",
#         "hi", "hey", "help", "how does", "what does", "bva",
#         "boundary", "equivalence", "what are"
#     ]

#     is_casual = any(keyword in last_message for keyword in casual_keywords)

#     if is_casual:
#         system_prompt = """
# You are Testron — an AI QA Test Case Quality Scoring Engine built to help QA engineers.

# You can:
# - Answer general QA questions (BVA, EP, test types, etc.)
# - Explain testing concepts clearly
# - Help with test case writing
# - Analyze test cases when provided

# Be helpful, friendly, and concise.
# When asked about yourself: You are Testron, a QA AI assistant.
# """
#     else:
#         system_prompt = TESTRON_PROMPT  # strict format for actual test cases

#     def generate():
#         try:
#             stream = client.chat.completions.create(
#                 model=MODEL,
#                 messages=[
#                     {"role": "system", "content": system_prompt},
#                     *messages
#                 ],
#                 stream=True,
#                 temperature=0,
#                 max_tokens=1024,
#             )

#             for chunk in stream:
#                 content = chunk.choices[0].delta.content
#                 if content:
#                     yield "data: " + json.dumps({"content": content}) + "\n\n"

#             yield "data: [DONE]\n\n"

#         except Exception as e:
#             yield "data: " + json.dumps({"error": str(e)}) + "\n\n"

#     return Response(
#         stream_with_context(generate()),
#         mimetype="text/event-stream",
#         headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
#     )


# # ── Batch Analyzer ─────────────────────────────────────────────
# @app.route("/batch-analyze", methods=["POST"])
# def batch_analyze():
#     data       = request.get_json()
#     test_cases = data.get("test_cases", [])
#     user_note  = data.get("note", "")

#     if not test_cases:
#         return jsonify({"error": "No test cases provided"}), 400

#     batch_text = f"Analyze this batch of {len(test_cases)} test cases:\n\n"

#     if user_note:
#         batch_text += f"User note: {user_note}\n\n"

#     for i, tc in enumerate(test_cases):
#         batch_text += (
#             f"TC-{i+1}\n"
#             f"  Title:          {tc.get('title', 'N/A')}\n"
#             f"  Preconditions:  {tc.get('pre', 'N/A')}\n"
#             f"  Steps:          {tc.get('steps', 'N/A')}\n"
#             f"  Test Data:      {tc.get('data', 'N/A')}\n"
#             f"  Expected:       {tc.get('expected', 'N/A')}\n\n"
#         )

#     def generate():
#         try:
#             stream = client.chat.completions.create(
#                 model=MODEL,
#                 messages=[
#                     {"role": "system", "content": TESTRON_PROMPT},
#                     {"role": "user",   "content": batch_text}
#                 ],
#                 stream=True,
#                 temperature=0,
#                 max_tokens=1024,
#             )

#             for chunk in stream:
#                 content = chunk.choices[0].delta.content
#                 if content:
#                     yield "data: " + json.dumps({"content": content}) + "\n\n"

#             yield "data: [DONE]\n\n"

#         except Exception as e:
#             yield "data: " + json.dumps({"error": str(e)}) + "\n\n"

#     return Response(
#         stream_with_context(generate()),
#         mimetype="text/event-stream",
#         headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
#     )


# # ── Excel Upload ───────────────────────────────────────────────
# @app.route("/upload-excel", methods=["POST"])
# def upload_excel():
#     if not EXCEL_SUPPORT:
#         return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

#     if "file" not in request.files:
#         return jsonify({"error": "No file uploaded"}), 400

#     file = request.files["file"]

#     if not file.filename.endswith((".xlsx", ".xls")):
#         return jsonify({"error": "Please upload an Excel file (.xlsx or .xls)"}), 400

#     try:
#         wb      = openpyxl.load_workbook(io.BytesIO(file.read()))
#         ws      = wb.active
#         headers = []
#         rows    = []

#         for i, row in enumerate(ws.iter_rows(values_only=True)):
#             if i == 0:
#                 headers = [str(c).strip() if c else "" for c in row]
#             else:
#                 if any(cell for cell in row):
#                     rows.append(dict(zip(headers, [str(c) if c is not None else "" for c in row])))

#         test_cases = []
#         for row in rows:
#             tc = {
#                 "id":       row.get("Test Case ID") or row.get("ID")    or row.get("id")    or "",
#                 "title":    row.get("Title")         or row.get("title") or row.get("Test Case Name") or "",
#                 "pre":      row.get("Preconditions") or row.get("preconditions") or "",
#                 "steps":    row.get("Steps to Reproduce") or row.get("Steps") or row.get("steps") or "",
#                 "data":     row.get("Test Data")     or row.get("test data") or "",
#                 "expected": row.get("Expected Result") or row.get("Expected Results") or row.get("expected result") or "",
#             }
#             if tc["title"] or tc["steps"]:
#                 test_cases.append(tc)

#         return jsonify({"test_cases": test_cases, "count": len(test_cases)})

#     except Exception as e:
#         return jsonify({"error": "Failed to read Excel: " + str(e)}), 500


# # ── Health Check ───────────────────────────────────────────────
# @app.route("/health")
# def health():
#     try:
#         if GROQ_API_KEY:
#             return jsonify({
#                 "status":   "ok",
#                 "provider": "Groq ⚡",
#                 "model":    MODEL
#             })
#         else:
#             return jsonify({"status": "error", "message": "GROQ_API_KEY not set"}), 503
#     except Exception as e:
#         return jsonify({"status": "error", "message": str(e)}), 503


# # ── Start Server ───────────────────────────────────────────────
# if __name__ == "__main__":
#     port = int(os.environ.get("PORT", 5050))
#     print("=" * 50)
#     print("  Testron AI Agent — Flask + Groq ⚡")
#     print(f"  Model      : {MODEL}")
#     print(f"  Running at : http://localhost:{port}")
#     print("=" * 50)
#     app.run(debug=True, port=5050)

#Groq Api Script V2 
from flask import Flask, render_template, request, Response, stream_with_context, jsonify
import json
import time  
import io
import os
from groq import Groq

# ── Load .env file ─────────────────────────────────────────────
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

app = Flask(__name__)

# ── CONFIG ─────────────────────────────────────────────────────
GROQ_API_KEY = os.environ.get("GROQ_API_KEY")
MODEL        = os.environ.get("MODEL", "llama-3.3-70b-versatile")
client       = Groq(api_key=GROQ_API_KEY)

# ── System Prompt ──────────────────────────────────────────────
TESTRON_PROMPT = """
You are Testron — an AI QA Test Case Quality Scoring Engine.

Your job is to evaluate SOFTWARE TEST CASE QUALITY.

INPUT:
You will receive a batch of test cases summarized in compact form.

CRITICAL RULES:
- NEVER repeat or quote test cases
- NEVER summarize the batch
- NEVER explain the test cases
- DO NOT write paragraphs
- ONLY produce the structured QA evaluation

Treat test case input as CONFIDENTIAL.

------------------------------------------------

Evaluate using professional QA metrics:

QUALITY CHECK
+ Title clarity
+ Preconditions present
+ Steps completeness
+ Expected result clarity
+ Test data presence

TEST DESIGN COVERAGE
+ Positive scenarios
+ Negative scenarios
+ Boundary Value Analysis (BVA)
+ Equivalence Partitioning (EP)
+ Edge cases

------------------------------------------------

You must also perform:

1. Detect duplicate test cases
2. Detect missing scenarios
3. Suggest Boundary Value test cases
4. Suggest Negative test cases
5. Score overall coverage

------------------------------------------------

SCORING MODEL (MAX 10)

+1 Title clarity
+1 Preconditions
+1 Steps completeness
+1 Expected result
+1 Test data
+1 Positive scenarios
+1 Negative scenarios
+1 Boundary value tests
+1 Equivalence partition tests
+1 Edge cases

------------------------------------------------

Respond ONLY in this format:

BATCH VERDICT: [X PASS | Y NEEDS FIX | Z CRITICAL] out of [total]

COMMON ISSUES:
- issue
- issue

DUPLICATE TEST CASES:
- TC-ID reason
- TC-ID reason

COVERAGE GAPS:
- missing technique
- missing scenario type

MISSING TEST CASES:
- missing scenario
- missing scenario

SUGGESTED BVA TESTS:
1. boundary case
2. boundary case

SUGGESTED NEGATIVE TESTS:
1. negative scenario
2. negative scenario

OVERALL SCORE: X/10

TOP 3 FIXES:
1. improvement
2. improvement
3. improvement

IMPORTANT:
If information is missing, infer what is likely missing based on QA best practices.
DO NOT output explanations.
"""

CASUAL_PROMPT = """
You are Testron — an AI QA Test Case Quality Scoring Engine built to help QA engineers.

You can:
- Answer general QA questions (BVA, EP, test types, etc.)
- Explain testing concepts clearly
- Help with test case writing
- Analyze test cases when provided

Be helpful, friendly, and concise.
When asked about yourself: You are Testron, a QA AI assistant.
"""


# ── Home Page ──────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")


# ── Chat Endpoint ──────────────────────────────────────────────
@app.route("/chat", methods=["POST"])
def chat():
    data     = request.get_json()
    messages = data.get("messages", [])

    last_message = messages[-1]["content"].lower() if messages else ""

    casual_keywords = [
        "what is", "what's", "explain", "define", "tell me",
        "who are you", "your name", "whats your", "hello",
        "hi", "hey", "help", "how does", "what does", "bva",
        "boundary", "equivalence", "what are"
    ]

    is_casual = any(keyword in last_message for keyword in casual_keywords)
    system_prompt = CASUAL_PROMPT if is_casual else TESTRON_PROMPT

    def generate():
        try:
            stream = client.chat.completions.create(
                model=MODEL,
                messages=[
                    {"role": "system", "content": system_prompt},
                    *messages
                ],
                stream=True,
                temperature=0,
                max_tokens=1024,
            )
            for chunk in stream:
                content = chunk.choices[0].delta.content
                if content:
                    yield "data: " + json.dumps({"content": content}) + "\n\n"
                    time.sleep(0.02)
            yield "data: [DONE]\n\n"
        except Exception as e:
            yield "data: " + json.dumps({"error": str(e)}) + "\n\n"

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ── Chat With File (ChatGPT style) ─────────────────────────────
@app.route("/chat-with-file", methods=["POST"])
def chat_with_file():
    message = request.form.get("message", "").strip()
    file    = request.files.get("file")

    file_context = ""

    if file and file.filename.endswith((".xlsx", ".xls")):
        try:
            wb      = openpyxl.load_workbook(io.BytesIO(file.read()))
            ws      = wb.active
            headers = []
            rows    = []

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c).strip() if c else "" for c in row]
                else:
                    if any(cell for cell in row):
                        rows.append(dict(zip(headers, [str(c) if c is not None else "" for c in row])))

            file_context = f"User uploaded an Excel file with {len(rows)} test cases:\n\n"

            for i, row in enumerate(rows):
                title    = row.get("Title")           or row.get("title")           or row.get("Test Case Name")  or "N/A"
                pre      = row.get("Preconditions")   or row.get("preconditions")   or "N/A"
                steps    = row.get("Steps")           or row.get("Steps to Reproduce") or row.get("steps")       or "N/A"
                data     = row.get("Test Data")       or row.get("test data")       or "N/A"
                expected = row.get("Expected Result") or row.get("Expected Results") or row.get("expected result") or "N/A"

                file_context += (
                    f"TC-{i+1}\n"
                    f"  Title:          {title}\n"
                    f"  Preconditions:  {pre}\n"
                    f"  Steps:          {steps}\n"
                    f"  Test Data:      {data}\n"
                    f"  Expected:       {expected}\n\n"
                )

        except Exception as e:
            file_context = f"Could not read Excel file: {str(e)}\n\n"

    # Combine file content + user message
    full_message = ""
    if file_context:
        full_message += file_context
    if message:
        full_message += f"User instruction: {message}"
    if not full_message:
        return jsonify({"error": "Please provide a message or file"}), 400

    # Use casual prompt if no file, strict if file uploaded
    system_prompt = TESTRON_PROMPT if file_context else CASUAL_PROMPT

    def generate():
        try:
            stream = client.chat.completions.create(
                model=MODEL,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user",   "content": full_message}
                ],
                stream=True,
                temperature=0,
                max_tokens=2048,
            )
            for chunk in stream:
                content = chunk.choices[0].delta.content
                if content:
                    yield "data: " + json.dumps({"content": content}) + "\n\n"
                    time.sleep(0.02)
            yield "data: [DONE]\n\n"
        except Exception as e:
            yield "data: " + json.dumps({"error": str(e)}) + "\n\n"

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ── Batch Analyzer ─────────────────────────────────────────────
@app.route("/batch-analyze", methods=["POST"])
def batch_analyze():
    data       = request.get_json()
    test_cases = data.get("test_cases", [])
    user_note  = data.get("note", "")

    if not test_cases:
        return jsonify({"error": "No test cases provided"}), 400

    batch_text = f"Analyze this batch of {len(test_cases)} test cases:\n\n"

    if user_note:
        batch_text += f"User note: {user_note}\n\n"

    for i, tc in enumerate(test_cases):
        batch_text += (
            f"TC-{i+1}\n"
            f"  Title:          {tc.get('title', 'N/A')}\n"
            f"  Preconditions:  {tc.get('pre', 'N/A')}\n"
            f"  Steps:          {tc.get('steps', 'N/A')}\n"
            f"  Test Data:      {tc.get('data', 'N/A')}\n"
            f"  Expected:       {tc.get('expected', 'N/A')}\n\n"
        )

    def generate():
        try:
            stream = client.chat.completions.create(
                model=MODEL,
                messages=[
                    {"role": "system", "content": TESTRON_PROMPT},
                    {"role": "user",   "content": batch_text}
                ],
                stream=True,
                temperature=0,
                max_tokens=1024,
            )
            for chunk in stream:
                content = chunk.choices[0].delta.content
                if content:
                    yield "data: " + json.dumps({"content": content}) + "\n\n"
            yield "data: [DONE]\n\n"
        except Exception as e:
            yield "data: " + json.dumps({"error": str(e)}) + "\n\n"

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ── Excel Upload ───────────────────────────────────────────────
@app.route("/upload-excel", methods=["POST"])
def upload_excel():
    if not EXCEL_SUPPORT:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]

    if not file.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": "Please upload an Excel file (.xlsx or .xls)"}), 400

    try:
        wb      = openpyxl.load_workbook(io.BytesIO(file.read()))
        ws      = wb.active
        headers = []
        rows    = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c).strip() if c else "" for c in row]
            else:
                if any(cell for cell in row):
                    rows.append(dict(zip(headers, [str(c) if c is not None else "" for c in row])))

        test_cases = []
        for row in rows:
            tc = {
                "id":       row.get("Test Case ID") or row.get("ID")    or row.get("id")    or "",
                "title":    row.get("Title")         or row.get("title") or row.get("Test Case Name") or "",
                "pre":      row.get("Preconditions") or row.get("preconditions") or "",
                "steps":    row.get("Steps to Reproduce") or row.get("Steps") or row.get("steps") or "",
                "data":     row.get("Test Data")     or row.get("test data") or "",
                "expected": row.get("Expected Result") or row.get("Expected Results") or row.get("expected result") or "",
            }
            if tc["title"] or tc["steps"]:
                test_cases.append(tc)

        return jsonify({"test_cases": test_cases, "count": len(test_cases)})

    except Exception as e:
        return jsonify({"error": "Failed to read Excel: " + str(e)}), 500


# ── Health Check ───────────────────────────────────────────────
@app.route("/health")
def health():
    try:
        if GROQ_API_KEY:
            return jsonify({
                "status":   "ok",
                "provider": "Groq ⚡",
                "model":    MODEL
            })
        else:
            return jsonify({"status": "error", "message": "GROQ_API_KEY not set"}), 503
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 503


# ── Start Server ───────────────────────────────────────────────
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5050))
    print("=" * 50)
    print("  Testron AI Agent — Flask + Groq ⚡")
    print(f"  Model      : {MODEL}")
    print(f"  Running at : http://localhost:{port}")
    print("=" * 50)
    app.run(debug=True, port=5050)